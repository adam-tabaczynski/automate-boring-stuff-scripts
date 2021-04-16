#! python3
# spreadsheet_cell_inverter.py - program that inverts rows and columns,
# f.i - moves value of cell A3 to C1, B7 to G2, etc.

import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('./chapter_13/example_tablo.xlsx')
sheet = wb.active

# copy all cells values to dictionary? to not overwrite?
# get max_row max_column? if C1 value = A3 value, the C1 value is gone
# in each iteration, get value of C1, save it to temp,
# C1 = A3; A3 = temp

max_row = sheet.max_row + 1
max_col = sheet.max_column + 1

inverted_cells_set = set()

for row in range(1, max_row):
  for column in range(1, max_col):
    cell_addr = get_column_letter(column) + str(row)
    inv_cell_addr =  get_column_letter(row) + str(column)
    inverted_cells_set.add(inv_cell_addr)
    
    if cell_addr in inverted_cells_set and row != column:
      continue
    
    temp = sheet[cell_addr].value
    sheet[cell_addr] = sheet[inv_cell_addr].value
    sheet[inv_cell_addr] = temp
  
    
wb.save('./chapter_13/inverted_table.xlsx')
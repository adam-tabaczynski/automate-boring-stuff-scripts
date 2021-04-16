#! python3
# spreadsheet+to_txt_files.py - program that put content of 
# .xlsx file into .txt files - each row will be a new line

from pathlib import Path
import os
import openpyxl
from openpyxl.utils import get_column_letter

files_directory = (Path.cwd() / 'chapter_13\\txt_files_from_spreadsheet')
files_directory = os.path.abspath(files_directory)
wb = openpyxl.load_workbook('./chapter_13/spreadsheet_from_txts.xlsx')
sheet = wb.active

max_row = sheet.max_row + 1
max_col = sheet.max_column + 1

for column in range(1, max_col):
  # creates a filename for each column
  filename = 'chapter_13\\txt_files_from_spreadsheet\\text_' + str(column) + '_from_xlsx.txt'
  txt_file_path = (Path.cwd() / filename)
  # if file already exist, delete it. Append mode will add new data at the end
  # instead of overwriting existing lines.
  if os.path.exists(txt_file_path):
    os.remove(txt_file_path)
    
  with open(txt_file_path, 'a') as file:
    for row in range(1, max_row):
      cell_addr = get_column_letter(column) + str(row)
      # skip cell if its empty.
      if(sheet[cell_addr].value == None):
        continue
      
      line = sheet[cell_addr].value
      file.write(line)
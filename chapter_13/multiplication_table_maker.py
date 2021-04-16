import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import sys

if len(sys.argv) != 2:
  raise Exception("Too many arguments!")


user_input = sys.argv[1]
number = int(user_input)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Multiplication table'
bold_font = Font(bold=True)

for row in range(1, number+2):
  for column in range(1, number+2):
    cell_address = get_column_letter(column) + str(row)
    if column == 1 and row == 1:
      pass
    elif row == 1:
      sheet[cell_address] = row * (column - 1)
      sheet[cell_address].font = bold_font
    elif column == 1:
      sheet[cell_address] = (row - 1) * column
      sheet[cell_address].font = bold_font
    else:
      sheet[cell_address] = (row - 1) * (column - 1)
    
wb.save('./chapter_13/multi_table.xlsx')
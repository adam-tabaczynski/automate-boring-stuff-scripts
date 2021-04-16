#! python3
# txt_files_to_spreadsheet.py - program that takes content of 
# several text files and insert each line into a file, one
# line of text per row. Contents of first file go to column A,
# contents of second file go to column B, etc.

from pathlib import Path
import os
import openpyxl
from openpyxl.utils import get_column_letter

files_directory = (Path.cwd() / 'chapter_13\\txt_files_for_spreadsheet')
files_directory = os.path.abspath(files_directory)
wb = openpyxl.Workbook()
sheet = wb.active

for foldername, subfolders, filenames in os.walk(files_directory):
  for index_file, filename in enumerate(filenames):
    file_path = (files_directory + '\\' + filename)
    with open(file_path) as file:
      for index_line, line in enumerate(file.readlines()):
          cell_addr = get_column_letter(index_file+1) + str(index_line+1)
          sheet[cell_addr] = line
          
wb.save('./chapter_13/spreadsheet_from_txts.xlsx')
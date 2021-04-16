import openpyxl
from openpyxl.utils import get_column_letter
import sys


if len(sys.argv) != 4:
  raise Exception("Type in starting row, number of blank rows, address of an excel file.")

starting_row = sys.argv[1]
starting_row = int(starting_row)

blank_rows = sys.argv[2]
blank_rows = int(blank_rows)

file_address = sys.argv[3]

wb = openpyxl.load_workbook(file_address)
sheet = wb.active

for row in range(sheet.max_row, starting_row-1, -1):
  for column in range(1, sheet.max_column+1):
    current_cell_address = get_column_letter(column) + str(row)
    target_cell_adress = get_column_letter(column) + str(row + blank_rows)
    sheet[target_cell_adress] = sheet[current_cell_address].value
    
for row in range(starting_row, starting_row + blank_rows):
  for column in range(1, sheet.max_column+1):
     current_cell_address = get_column_letter(column) + str(row)
     sheet[current_cell_address] = ''
     
     
wb.save('./chapter_13/blank_row_table.xlsx')